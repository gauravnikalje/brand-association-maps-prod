import openpyxl
wb = openpyxl.load_workbook(r"BAM_output_52a4c94d.xlsx", read_only=True)
for name in wb.sheetnames:
    ws = wb[name]
    rows = list(ws.iter_rows(max_row=4, values_only=True))
    print(f"\n=== {name} ({ws.max_row} rows x {ws.max_column} cols) ===")
    for r in rows:
        print(list(r))
wb.close()
